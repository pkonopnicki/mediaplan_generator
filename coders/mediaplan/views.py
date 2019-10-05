from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.views import View
import openpyxl
from openpyxl import load_workbook
import itertools
import pandas as pd
import xlwt
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Color, PatternFill, Font, Border
from .column_order import column_order


# Create your views here.


def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    elif 'excel_upload' in request.POST:
        excel_file = request.FILES["excel_file"]

        excel_data = pd.read_excel(excel_file, sheetname=0)

        column_names = excel_data.columns.tolist()
        vendor_column = excel_data[column_names[0]].dropna().tolist()
        ad_serving_type = excel_data['AdServingType'].dropna().tolist()

        vendor_len = len(vendor_column)
        column_len = len(column_names)
        col_n = '_'.join(str(e) for e in column_names)

        columns_list = []
        for i in range(1, len(column_names)):
            columns_dictionary = {
                "column_name": excel_data.columns[i],
                "column_options": excel_data[column_names[i]].dropna().tolist()
            }
            columns_list.append(columns_dictionary)

        vendor_list = []
        for i in range(0, len(vendor_column)):
            vendor_dictionary = {
                "vendor_id": str(i),
                "vendor_name": vendor_column[i],
                "option_list": columns_list
            }
            vendor_list.append(vendor_dictionary)

        ctx = {
            'vendor_list': vendor_list,
            'vendor_len': vendor_len,
            'column_len': column_len,
            'column_names': col_n,
            "vendor_name": vendor_column,
            'ad_serving_type': ad_serving_type
        }

        return render(request, 'redirect.html', ctx)

    elif 'go_next' in request.POST:

        vendor_count = request.POST.get("vendor count")
        column_count = request.POST.get("column count")
        column_names_concat = request.POST.get("column names")
        column_names = column_names_concat.split("_")

        vendor_list = []
        for j in range(0, int(vendor_count)):
            columns_list = []
            for i in range(1, int(column_count)):
                columns_dictionary = {
                    "column_name": column_names[i],
                    "column_options": request.POST.getlist("{} {}".format(column_names[i], str(j)))
                }
                columns_list.append(columns_dictionary)

            vendor_dictionary = {
                "vendor_name": request.POST.get("vendor {}".format(str(j))),
                "option_list": columns_list
            }
            vendor_list.append(vendor_dictionary)

        to_concat = []
        for one_vendor_dict in vendor_list:
            new_option_dict = {}
            for option_dict in one_vendor_dict['option_list']:
                column_name, option_value = None, None
                # get column name and column values
                for k, v in option_dict.items():
                    if 'name' in k:
                        column_name = v
                    if 'options' in k:
                        option_value = v
                if column_name and option_value:
                    new_option_dict[column_name] = option_value

            # put all list to same length in order to build a dataframe.
            max_length = max([len(v) for v in new_option_dict.values()])
            for k, v in new_option_dict.items():
                if len(v) < max_length:
                    new_option_dict.update({k: v + [None] * (max_length - len(v))})
            # add the vendor column
            new_option_dict.update({'Vendor': [one_vendor_dict['vendor_name']] * max_length})
            # create a dataframe for this vendor
            to_concat.append(pd.DataFrame(new_option_dict))
        df = pd.concat(to_concat).reset_index(drop=True)
        cols = df.columns.tolist()
        cols = cols[-1:] + cols[:-1]
        df = df[cols]

        output_file = 'MediaPlan_output.xlsx'
        global output_df
        output_df = pd.DataFrame()

        vendor_names = df["Vendor"].dropna().unique().tolist()

        # buy model buy rate df and add kpi and sov df

        buymodel = request.POST.getlist("buymodel")
        buyrate = request.POST.getlist("buyrate")
        sov = request.POST.getlist("sov")
        kpi = request.POST.getlist("kpi")


        buymodel_df = pd.DataFrame()

        buymodel_df["Vendor"] = vendor_names
        buymodel_df["Buy Model"] = buymodel
        buymodel_df["Buy Rate"] = buyrate
        buymodel_df["Planned SOV %"] = sov
        buymodel_df["KPI (If Needed)"] = kpi


        # planned net cost df

        plannednetcost = request.POST.getlist("plannednetcost")
        plannedimpressionsmultiplier = request.POST.getlist("plannedimpressions")

        plannednetcost_df = pd.DataFrame()

        plannednetcost_df["Vendor"] = vendor_names
        plannednetcost_df["Planned Net Cost"] = plannednetcost
        plannednetcost_df["Planned Impressions Multiplier"] = plannedimpressionsmultiplier

        # ad serving and reporting fee


        serving_types = request.POST.getlist("serving")
        adrate = request.POST.getlist("adrate")
        reportingfee = request.POST.getlist("reportingrate")
        adverificationrate = request.POST.getlist("adverificationrate")


        serving_df = pd.DataFrame()

        serving_df["AdServingType"] = serving_types
        serving_df["Ad Serving Rate"] = adrate
        serving_df["Reporting Fee Rate"] = reportingfee
        serving_df["Ad Verification Rate"] = adverificationrate



        # Creating combinations of the input variables

        def give_all_combinations():
            first_col = df.columns[0]
            first_col_unique = df[first_col].unique()
            global output_df
            for val in first_col_unique:
                df1 = df.loc[df[first_col] == val]

                value_list = [df1[x].dropna().unique() for x in df1.columns]

                data = [e for e in itertools.product(*value_list)]

                df1 = pd.DataFrame(data, columns=df1.columns)
                output_df = output_df.append(df1, ignore_index=True)

        # Adding buy model

        def add_buymodel():
            global output_df
            df_buymodel_dictionary = buymodel_df
            output_df['Vendor_left'] = output_df['Vendor']
            output_df['Vendor_left'] = output_df['Vendor_left'].str.lower()
            output_df['Vendor_left'] = output_df['Vendor_left'].str.replace(' ', '')
            df_buymodel_dictionary['Vendor_right'] = df_buymodel_dictionary['Vendor']
            df_buymodel_dictionary['Vendor_right'] = df_buymodel_dictionary['Vendor_right'].str.lower()
            df_buymodel_dictionary['Vendor_right'] = df_buymodel_dictionary['Vendor_right'].str.replace(' ', '')
            output_df = pd.merge(output_df, df_buymodel_dictionary, left_on='Vendor_left',
                                 right_on='Vendor_right', suffixes=('', '_right_join'), how='left')
            output_df = output_df.drop(columns=['Vendor_left', 'Vendor_right_join', 'Vendor_right'])

        # Adding adserving fee

        def add_serving():
            global output_df
            df_serving_dictionary = serving_df
            output_df['AdServingType_left'] = output_df['AdServingType']
            output_df['AdServingType_left'] = output_df['AdServingType_left'].str.lower()
            output_df['AdServingType_left'] = output_df['AdServingType_left'].str.replace(' ', '')
            df_serving_dictionary['AdServingType_right'] = df_serving_dictionary['AdServingType']
            df_serving_dictionary['AdServingType_right'] = df_serving_dictionary['AdServingType_right'].str.lower()
            df_serving_dictionary['AdServingType_right'] = df_serving_dictionary['AdServingType_right'].str.replace(' ',
                                                                                                                    '')
            output_df = pd.merge(output_df, df_serving_dictionary, left_on='AdServingType_left',
                                 right_on='AdServingType_right', suffixes=('', '_right_join'), how='left')
            output_df = output_df.drop(
                columns=['AdServingType_left', 'AdServingType_right_join', 'AdServingType_right'])

        # Adding budget with multiple granularity option

        def add_budget():
            global output_df
            df_budget_dictionary = plannednetcost_df

            list_basic = ['Vendor']
            list_left = [el + "_left" for el in list_basic]
            list_right = [el + "_right" for el in list_basic]
            list_join_right = [el + "_join" for el in list_right]
            list_join = list_left + list_right + list_join_right

            output_df[list_left] = output_df[list_basic]
            df_budget_dictionary[list_right] = df_budget_dictionary[list_basic]
            output_df = pd.merge(output_df, df_budget_dictionary, left_on=list_left,
                                 right_on=list_right, suffixes=('', '_right_join'), how='left')
            output_df = output_df.drop(columns=list_join)
            output_df.loc[output_df.duplicated(subset=list_basic), "Planned Net Cost"] = 0
            output_df.loc[output_df.duplicated(subset=list_basic), "Planned Impressions Multiplier"] = 0

        # use all methods

        give_all_combinations()
        add_buymodel()
        add_budget()
        add_serving()

        # get all columns from the form and add them to df

        client_name = request.POST.get("client_name")
        campaign_description = request.POST.get("campaign_description")
        franchise_name = request.POST.get("franchise_name")
        campaign_type = request.POST.get("campaign_type")
        product_name = request.POST.get("product_name")
        product_detail = request.POST.get("product_detail")
        campaign_timing = request.POST.get("campaign_timing")
        year = request.POST.get("year")
        campaign_region = request.POST.get("campaign_region")
        campaign_id = request.POST.get("campaign_id")
        agency_fee_rate = request.POST.get("agency_fee_rate")
        verification_buffer_amount = request.POST.get("verification_buffer_amount")
        service_fee_rate = request.POST.get("service_fee_rate")
        adserving_buffer_amount = request.POST.get("adserving_buffer_amount")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        placement_phase = request.POST.get("placement_phase")
        placement_objective = request.POST.get("placement_objective")


        output_df["Client Name"] = client_name
        output_df["Campaign Description"] = campaign_description
        output_df["Franchise Name"] = franchise_name
        output_df["Campaign Type"] = campaign_type
        output_df["Product Name"] = product_name
        output_df["Product Detail"] = product_detail
        output_df["Campaign Timing"] = campaign_timing
        output_df["Year"] = year
        output_df["Campaign Region"] = campaign_region
        output_df["Campaign ID"] = campaign_id
        output_df["Placement Phase (If Needed)"] = placement_phase
        output_df["Placement Objective (If Needed)"] = placement_objective
        output_df["Agency Fee Rate"] = agency_fee_rate
        output_df["Verification Buffer Amount"] = verification_buffer_amount
        output_df["Service Fee Rate"] = service_fee_rate
        output_df["Ad Serving Buffer Amount"] = adserving_buffer_amount

        try:
            output_df["Start Date"] = start_date
        except:
            output_df["Start Date"] = None
        try:
            output_df["End Date"] = end_date
        except:
            output_df["End Date"] = None

        # add empty columns as per mediaplan

        output_df["DCM Link"] = None
        output_df["Length"] = None
        output_df["Twitter Card Name"] = None
        output_df["Clickthrough UTM"] = None
        output_df["Base URL"] = None
        output_df["UTM_Source"] = None
        output_df["Source"] = None
        output_df["UTM_Medium"] = None
        output_df["Medium"] = None
        output_df["UTM_Term"] = None
        output_df["Term"] = None
        output_df["UTM_Content"] = None
        output_df["Content"] = None
        output_df["UTM_Campaign"] = None
        output_df["Campaign"] = None

        # add reporting source (v or a) depending on vendor

        selfserved = ["Twitter", "Snapchat", "Reddit", "TTD", "TradeDesk", "Trade Desk",
                      "The Trade Desk", "TheTradeDesk", "Google", "Google SEM", "SEM",
                      "GDN", "Search", "Google Search", "YT", "YouTube", "Youtube", "Adwords",
                      "Google Adwords", "FB", "IG", "Facebook", "Instagram", "FB&IG",
                      "FB & IG", "Facebook Performance", "Facebook Branding", "Instagram Performance", "Instagram Branding",
                      "FB & IG Branding", "FB & IG Performance", "FB&IG Branding", "FB&IG Performance"]

        selfserved_lower = [x.lower() for x in selfserved]
        selfserved_upper = [x.upper() for x in selfserved]
        selfserved_combined = selfserved + selfserved_lower + selfserved_upper

        output_df["Reporting Source"] = ["V" if el in selfserved_combined else "A" for el in output_df["Vendor"]]

        # add formulas

        campaign_name = '=CONCATENATE(Table1[@[Franchise Name]],"_",Table1[@[Campaign Type]],"_",Table1[@[Product Name]],"_",Table1[@[Campaign Timing]],"_",Table1[@[Year]],"_",Table1[@[Campaign Region]])'
        planned_impressions = '=Table1[@[Planned Net Cost]]/Table1[@[Planned Impressions Multiplier]]*1000'
        planned_units = '=Table1[@[Planned Net Cost]]/Table1[@[CPM / Cost Per Unit]]'
        agency_fee_cost = '=Table1[@[Agency Fee Rate]]*Table1[@[Planned Net Cost]]'
        ad_verification_cost = '=Table1[@[Ad Verification Rate]]*(Table1[@[Planned Impressions]]/1000)'
        ad_serving_cost = '=Table1[@[Ad Serving Rate]]*(Table1[@[Planned Impressions]]/1000)'
        verification_buffer_total = '=Table1[@[Ad Verification Cost]]*(Table1[@[Verification Buffer Amount]])'
        reporting_fee_cost = '=Table1[@[Reporting Fee Rate]]*(Table1[@[Planned Impressions]]/1000)'
        service_fee_cost = '=(Table1[@[Ad Serving Cost]]+Table1[@[Ad Server Buffer Total]])*Table1[@[Service Fee Rate]]'
        adserver_buffer_total = '=Table1[@[Ad Serving Buffer Amount]]*Table1[@[Ad Serving Cost]]'
        total_cost = '=SUM(Table1[@[Ad Server Buffer Total]],Table1[@[Reporting Fee Cost]],Table1[@[Service Fee Cost]],Table1[@[Ad Serving Cost]],Table1[@[Ad Verification Cost]],Table1[@[Verification Buffer Total]],Table1[@[Agency Fee Cost]],Table1[@[Planned Net Cost]])'
        placement_name = '=IF(Table1[@[Ad Size (WxH)]]="PKG","",CONCATENATE(Table1[@[Campaign ID]],"_",Table1[@[Partner Name]],"_' \
                         '",Table1[@[Country]],"_",Table1[@[Targeting]],"_",Table1[@[Creative (If Needed)]],"_",Table1[@[Copy (If Needed)]],' \
                         '"_",Table1[@[Buy Model]],"_",Table1[@[CPM / Cost Per Unit]],"_",Table1[@[Start Date]],"_",Table1[@[Ad Serving Type]],' \
                         '"_",Table1[@[Reporting Fee Rate]],"_",Table1[@[KPI (If Needed)]],"_",Table1[@[Placement Objective (If Needed)]],"_",Table1[@[Placement Phase (If Needed)]],"_",'\
                         'Table1[@[Service Fee Rate]],"_",Table1[@[Ad Verification Rate]],"_",Table1[@[Reporting Source]],"_",Table1[@[Device]],"_",Table1[@[Ad Size (WxH)]],"_",Table1[@[Ad Type]],"_",' \
                         'Table1[@[Placement Description]],"_",Table1[@[Package Description]]))'
        self_serve_campaign_name = '=CONCATENATE(Table1[@[Campaign ID]],"_",Table1[@[Campaign Type]],"_",Table1[@[Partner Name]],"_",Table1[@[Country]],"_",Table1[@[Creative (If Needed)]])'


        # change column order

        output_df = output_df.rename(columns={"Buy Rate": "CPM / Cost Per Unit",
                                              "Vendor": "Partner Name",
                                              "PackageDescription": "Package Description",
                                              "PlacementDescription": "Placement Description",
                                              "AdSize": "Ad Size (WxH)",
                                              "AdType": "Ad Type",
                                              "AdServingType": "Ad Serving Type",
                                              "Copy": "Copy (If Needed)",
                                              "Creative": "Creative (If Needed)",
                                              })

        # add excel table

        workbook = xlsxwriter.Workbook(output_file, {'nan_inf_to_errors': True})
        worksheet = workbook.add_worksheet("mediaplan")

        '''output_df = output_df[column_order]'''
        total_rows = len(output_df.index)+10

        worksheet.add_table('A1:BO{}'.format(total_rows), {'data': output_df.values.tolist(),
                                         'columns': [{'header': c} for c in output_df.columns.tolist()] +
                                                    [{'header': 'Campaign Name',
                                                      'formula': campaign_name}
                                                     ] +
                                                    [{'header': 'Planned Units (eg. CPV, CPE, CPI)',
                                                      'formula': planned_units}
                                                     ] +
                                                    [{'header': 'Agency Fee Cost',
                                                      'formula': agency_fee_cost}
                                                     ] +
                                                    [{'header': 'Ad Verification Cost',
                                                      'formula': ad_verification_cost}
                                                     ] +
                                                    [{'header': 'Verification Buffer Total',
                                                      'formula': verification_buffer_total}
                                                     ] +
                                                    [{'header': 'Reporting Fee Cost',
                                                      'formula': reporting_fee_cost}
                                                     ] +
                                                    [{'header': 'Service Fee Cost',
                                                      'formula': service_fee_cost}
                                                     ] +
                                                    [{'header': 'Ad Server Buffer Total',
                                                      'formula': adserver_buffer_total}
                                                     ] +
                                                    [{'header': 'Total Cost',
                                                      'formula': total_cost}
                                                     ] +
                                                    [{'header': 'Placement Name',
                                                      'formula': placement_name}
                                                     ] +
                                                    [{'header': 'Self Serve Campaign Name',
                                                      'formula': self_serve_campaign_name}
                                                     ] +
                                                    [{'header': 'Planned Impressions',
                                                      'formula': planned_impressions}
                                                     ] +
                                                    [{'header': 'Ad Serving Cost',
                                                      'formula': ad_serving_cost}
                                                     ]
            ,
                                         'style': 'Table Style Medium 9',
                                         })

        workbook.close()



        # adding background and font color

        #whiteFont = Font(color='FFFFFF')
        #blackFill = PatternFill(bgColor='000000', fill_type='solid')


        wb = load_workbook(output_file)
        ws = wb.active
        #ws['A2'].fill = blackFill
        #ws['A2'].font = whiteFont


        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="MediaPlan_output.xls"'
        wb.save(response)


        return response
